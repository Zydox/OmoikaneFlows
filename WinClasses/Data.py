# Copyright (c) 2025-2026 Erik Persson
# Licensed for free use only. Copying, modification, or redistribution prohibited.

import builtins
from Functions.Common import *
from Functions.DataHandling import *
for name in dir ():
    if not name.startswith ("_"):
        builtins.__dict__[name] = globals ()[name]
import re
builtins.re = re
import os
builtins.os = os
import copy
builtins.copy = copy
import openpyxl
builtins.openpyxl = openpyxl
import threading
builtins.threading = threading

class Data:
	def __init__ (self, cls_steps):
		self.cls_steps = cls_steps
		self.objects = self.cls_steps.objects
		self.objects['Lock']['Function.Variable.Change'] = threading.RLock ()

	
	def dynamic_function (self, function, to_class = True):
		if to_class is True:
			func = {}
			exec (function['Code'], func)
			setattr (Data, function['Function'], func[function['Function']])
		else:
			exec (function['Code'], globals ())
			builtins.__dict__[function['Function']] = globals ()[function['Function']]

	
	def execute (self, isolation = None, **opts):
		try:
			if opts['Function'] == 'DATA:Verify':
				return self.step_verify (isolation, **opts)
			elif opts['Function'] == 'DATA:Convert':
				return self.step_convert (isolation, **opts)
			elif opts['Function'] == 'DATA:Replace':
				return self.step_replace (isolation, **opts)
			elif opts['Function'] == 'DATA:Calculate':
				return self.step_calculate (isolation, **opts)
			elif opts['Function'] == 'DATA:Transform':
				return self.step_transform (isolation, **opts)
			elif opts['Function'] == 'DATA:Date':
				return self.step_date (isolation, **opts)
			elif opts['Function'] == 'DATA:Variable.Set':
				return self.step_variable_set (isolation, **opts)
			elif opts['Function'] == 'DATA:Variable.Change':
				return self.step_variable_change (isolation, **opts)
			elif opts['Function'] == 'DATA:File.Read':
				return self.step_file_read (isolation, **opts)
			elif opts['Function'] == 'DATA:File.Create':
				return self.step_file_create (isolation, **opts)
			elif opts['Function'] == 'DATA:SAP.UnconvertedTableToList':
				return self.step_sap_unconverted_table_to_list (isolation, **opts)
			elif opts['Function'] == 'DATA:List.':
				return self.step_list (isolation, **opts)
			else:
				return False, {'Status':'WARNING', 'Title':'DATA:execute: Aborted', 'Message':'Unknown function:' + str (opts)}
		except Exception as e:
			st = internal_zyd_stacktrace (e, FunctionOpts=opts)
			return False, {'Status':'CRASH', 'Title':'DATA:execute: Crashed', 'Message':'Crashed with the error: ' + str (e) + ' for step:\n' + str (opts), 'StackTrace':st}

	
	def step_verify (self, isolation, **opts):
		if 'Value' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_verify: Aborted', 'Message':'A Value is required for this function.'}
		if 'Verify' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_verify: Aborted', 'Message':'A Verify rule is required for this function.'}
		if not isinstance (opts['Verify'], dict):
			return False, {'Status':'WARNING', 'Title':'DATA:step_verify: Aborted', 'Message':'The Verify rule needs to be a dict.'}
		result = None
		if 'Convert' in opts:
			if not isinstance (opts['Convert'], list):
				return False, {'Status':'WARNING', 'Title':'DATA:step_verify: Aborted', 'Message':'The Convert option needs to be a list of conversion rules.'}
			value = copy.deepcopy (opts['Value'])
			for convert in opts['Convert']:
				status, value = datahandling_convert (value, **convert)
				if status is not True:
					break
		else:
			value = opts['Value']
			status = True
		
		if status is True:
			status, result = datahandling_verify (value, **opts['Verify'])
		
		if 'ReturnVariable' in opts and isinstance (opts['ReturnVariable'], str):
			self.cls_steps.internal_value_set_value (isolation, 'Variable', status, opts['ReturnVariable'])
		if 'ReturnVariable.Global' in opts and isinstance (opts['ReturnVariable.Global'], str):
			self.cls_steps.internal_value_set_value (None, 'Variable', status, opts['ReturnVariable.Global'])
		
		if status is True and 'OnTrue' in opts:
			self.cls_steps.execute_steps (opts['OnTrue'], isolation)
		elif status is False and result is not None and (('Aborted' in result and result['Aborted'] is True) or result['Status'] not in ['OK', 'WARNING']):
			return status, result
		elif status is False and 'OnFalse' in opts:
			self.cls_steps.execute_steps (opts['OnFalse'], isolation)
		return True, {'Status':'OK'}
	
	
	def step_file_read (self, isolation, **opts):
		result = self.internal_read_file (**opts)
		if opts['Type'] == 'JSON':
			self.cls_steps.internal_value_set_value (isolation, 'Variable', result, opts['ReturnVariable.Data'])
			return True, {'Status':'OK', 'Data':result}
		elif opts['Type'] == 'CSV':
			if isinstance (result, list):
				headers = []
				data = result
				if 'HeadersFromFile' in opts and opts['HeadersFromFile'] is True:
					headers = data[0]
					data = data[1:]
				if 'HeadersVerify' in opts and isinstance (opts['HeadersVerify'], list):
					if opts['HeadersVerify'] != headers:
						return False, {'Status':'WARNING', 'Title':'', 'Message':''}
				
				self.cls_steps.internal_value_set_value (isolation, 'Variable', data, opts['ReturnVariable.Data'])
				status = True
				
				return True, {'Status':'OK', 'Headers':headers, 'Data':data}
		elif opts['Type'] == 'Binary':
			self.cls_steps.internal_value_set_value (isolation, 'Variable', result, opts['ReturnVariable.Data'])
			return True, {'Status':'OK', 'Message':str (len (result)) + ' bytes read.'}
		
	
	def step_file_create (self, isolation, **opts):
		if 'File' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_file_create: Aborted', 'Message':'No File was provided for the function.'}
		if 'Type' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_file_create: Aborted', 'Message':'No Type was provided for the function.'}
		if 'Data' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_file_create: Aborted', 'Message':'No Data was provided for the function.'}
		
		if opts['Type'] == 'JSON':
			with open (opts['File'], 'w') as file:
				status, JSON = zyd_json (opts['Data'], Encode=True)
				file.write (JSON)
		elif opts['Type'] == 'Binary':
			with open (opts['File'], 'wb') as file:
				file.write (opts['Data'])
		elif opts['Type'] == 'CSV':
			csv_opts = {}
			for option in [['quotechar', 'CSV.Quotechar'], ['delimiter', 'CSV.Delimiter']]:
				if option[1] in opts and isinstance (opts[option[1]], str):
					csv_opts[option[0]] = opts[option[1]]
			if 'CSV.QuoteRule' in opts:
				if opts['CSV.QuoteRule'] == 'All':
					csv_opts['quoting'] = csv.QUOTE_ALL
				elif opts['CSV.QuoteRule'] == 'Minimal':
					csv_opts['quoting'] = csv.QUOTE_MINIMAL
				elif opts['CSV.QuoteRule'] == 'NonNumeric':
					csv_opts['quoting'] = csv.QUOTE_NONNUMERIC
				elif opts['CSV.QuoteRule'] == 'None':
					csv_opts['quoting'] = csv.QUOTE_NONE
				elif opts['CSV.QuoteRule'] == 'Strings':
					csv_opts['quoting'] = csv.QUOTE_STRINGS
				else:
					return False, {'Status':'WARNING', 'Title':'DATA:step_file_create: Aborted', 'Message':'CSV.QuoteRule "' + str (opts['CSV.QuoteRule']) + '" is not valid.'}
			
			with open (opts['File'], 'w', newline='') as csvfile:
				csvwriter = csv.writer (csvfile, **csv_opts)
				if 'CSV.Headers' in opts:
					if isinstance (opts['CSV.Headers'], list):
						csvwriter.writerow (opts['CSV.Headers'])
					else:
						return False, {'Status':'WARNING', 'Title':'DATA:step_file_create: Aborted', 'Message':'CSV.Headers needs to be a list, not "' + str (type (opts['CSV.Headers'])) + '".'}
				csvwriter.writerows (opts['Data'])
			
		return True, {'Status':'OK'}
	
	
	def step_sap_unconverted_table_to_list (self, isolation, **opts):
		if 'File' in opts and isinstance (opts['File'], str):
			file_opts = {}
			if 'File.Encoding' in opts and isinstance (opts['File.Encoding'], str):
				file_opts['encoding'] = opts['File.Encoding']
			with open (opts['File'], 'r', **file_opts) as fp:
				data = fp.read ().replace ('\r','').split ('\n')
			fp.close ()
		elif 'Data' in opts and isinstance (opts['Data'], str):
			data = opts['Data'].replace ('\r','').split ('\n')
		else:
			return False, {'Status':'WARNING', 'Title':'DATA:step_sap_unconverted_table_to_list: Aborted', 'Message':'No input data was provided for the function.'}
		
		tableid = -1
		tables = []
		header = ''
		headers = []
		unknown = []
		debuglog = (True if 'Debug.Log' in opts and opts['Debug.Log'] is True else False)
		for irow in range (0, len (data)):
			line = data[irow]
			if len (line) <= 3:
				if debuglog:
					self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_sap_unconverted_table_to_list', 'TableID':tableid, 'Part':'SKIPPED_LINE', 'Line':line})
			elif line.count ('-') == len (line):
				if debuglog:
					self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_sap_unconverted_table_to_list', 'TableID':tableid, 'Part':'HEADER_LOGIC', 'Line':line})
				if len (data) > irow+2 and len (data[irow+1]) > 3 and len (data[irow+2]) > 3 and data[irow+1][0] + data[irow+1][-1] == '||' and data[irow+1] != header and (data[irow+2] == '|' + line[1:-1] + '|' or data[irow+2] == line) and (irow == 0 or data[irow-1] != header):
					header = data[irow+1]
					headers = []
					tableid += 1
					tables.append ([])
					if debuglog:
						self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_sap_unconverted_table_to_list', 'TableID':tableid, 'Part':'HEADER', 'Line':data[irow+1]})
					for ichar in range (0, len (header) - 1):
						if header[ichar] == '|':
							headers.append (['', ichar + 1, 0])
						elif len (headers) > 0:
							headers[len (headers) - 1][0] += header[ichar]
							headers[len (headers) - 1][2] += 1
					if debuglog:
						self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_sap_unconverted_table_to_list', 'TableID':tableid, 'Part':'HEADER_CAPTURED', 'Headers':headers})
					
					entry = []
					for value in headers:
						if 'HeadersTrim' in opts and opts['HeadersTrim'] is True:
							entry.append (value[0].strip ())
						else:
							entry.append (value[0])
					if 'ReturnVariable.Headers' in opts and isinstance (opts['ReturnVariable.Headers'], str):
						if 'OnlyTableNo' in opts:
							if opts['OnlyTableNo'] == tableid:
								self.cls_steps.internal_value_set_value (isolation, 'Variable', entry, opts['ReturnVariable.Headers'])
						elif tableid == 0:
							self.cls_steps.internal_value_set_value (isolation, 'Variable', [entry], opts['ReturnVariable.Headers'])
						else:
							entries = self.cls_steps.internal_value_get_value (isolation, 'Variable', opts['ReturnVariable.Headers'])
							entries.append (entry)
							self.cls_steps.internal_value_set_value (isolation, 'Variable', entries, opts['ReturnVariable.Headers'])
					if 'HeadersInData' in opts and opts['HeadersInData'] is True:
						tables[tableid].append (entry)
			elif line[0] + line[-1] == '||' and header != '' and line.count ('-') != len (line) - 2 and line != header:
				entry = []
				if debuglog:
					self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_sap_unconverted_table_to_list', 'TableID':tableid, 'Part':'DATA_CAPTURE', 'Line':line})
				for column in headers:
					entry.append (line[column[1]:column[1]+column[2]])
				tables[tableid].append (entry)
				if debuglog:
					self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_sap_unconverted_table_to_list', 'TableID':tableid, 'Part':'DATA_CAPTURED', 'Columns':entry})
			elif line != header and len (line) != line.count ('|') + line.count ('-'):
				if 'ReturnVariable.UnknownLines' in opts:
					unknown.append (line)
				if debuglog:
					self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_sap_unconverted_table_to_list', 'TableID':tableid, 'Part':'UNKNOWN_LINE', 'Line':line})
		
		if 'ReturnVariable.UnknownLines' in opts and isinstance (opts['ReturnVariable.UnknownLines'], str):
			self.cls_steps.internal_value_set_value (isolation, 'Variable', unknown, opts['ReturnVariable.UnknownLines'])
		
		if 'ReturnVariable.Data' in opts and isinstance (opts['ReturnVariable.Data'], str):
			if 'OnlyTableNo' not in opts:
				self.cls_steps.internal_value_set_value (isolation, 'Variable', tables, opts['ReturnVariable.Data'])
			elif len (tables) >= opts['OnlyTableNo'] + 1:
				self.cls_steps.internal_value_set_value (isolation, 'Variable', tables[opts['OnlyTableNo']], opts['ReturnVariable.Data'])
		
		if 'Debug.Tables.Console' in opts and opts['Debug.Tables.Console'] is True:
			for table in range (0, len (tables)):
				print (('\n' if table > 0 else '') + '===============================================================================================\n' + 'TABLE=' + str (table) + '\n===============================================================================================')
				for row in tables[table]:
					print (row)
		return True, {'Status':'OK'}
	
	
	def step_calculate (self, isolation, **opts):
		if 'Calculation' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_calculate: Aborted', 'Message':'No Calculation was provided.'}
		if not isinstance (opts ['Calculation'], str) and ('Calculation.ToString' not in opts or opts['Calculation.ToString'] is not True):
			return False, {'Status':'WARNING', 'Title':'DATA:step_calculate: Aborted', 'Message':'Calculation needs to be a string.'}
		if 'Calculation.Round' in opts and not isinstance (opts['Calculation.Round'], int):
			return False, {'Status':'WARNING', 'Title':'DATA:step_calculate: Aborted', 'Message':'Calculation.Round needs to be a integer.'}
		if 'Calculation.Ceil' in opts and not isinstance (opts['Calculation.Ceil'], int):
			return False, {'Status':'WARNING', 'Title':'DATA:step_calculate: Aborted', 'Message':'Calculation.Ceil needs to be a integer.'}
		if 'Calculation.Floor' in opts and not isinstance (opts['Calculation.Floor'], int):
			return False, {'Status':'WARNING', 'Title':'DATA:step_calculate: Aborted', 'Message':'Calculation.Floor needs to be a integer.'}
		
		calc = copy.deepcopy (opts['Calculation'])
		if 'Calculation.ToString' in opts and opts['Calculation.ToString'] is True:
			calc = str (calc)
		if 'Calculation.CleanOperators' in opts and opts['Calculation.CleanOperators'] is True:
			for operator in ['*', '/', '-', '+', '%']:
				while operator + operator in calc:
					calc = calc.replace (operator + operator, operator)
				if calc[-1] == operator:
					calc = calc[:-1]
		try:
			status, value = zyd_calc (calc)
			if status is not True:
				return False, {'Status':'WARNING', 'Title':'DATA:step_calculate: Aborted', 'Message':'Calculation aborted with the following error:' + str (value)}
			if 'Calculation.Round' in opts:
				value = round (value, opts['Calculation.Round'])
			if 'Calculation.Ceil' in opts:
				value = internal_zyd_ceil (value, Decimals=opts['Calculation.Ceil'])
			if 'Calculation.Floor' in opts:
				value = internal_zyd_floor (value, Decimals=opts['Calculation.Floor'])
		except:
			return False, {'Status':'CRASH', 'Title':'DATA:step_calculate: Crashed', 'Message':'Calculation for "' + str (opts['Calculation']) + '" ("' + str (calc) + '") crashed.'}
		
		if 'ReturnVariable' in opts and isinstance (opts['ReturnVariable'], str):
			self.cls_steps.internal_value_set_value (isolation, 'Variable', value, opts['ReturnVariable'])
		if 'ReturnVariable.Global' in opts and isinstance (opts['ReturnVariable.Global'], str):
			self.cls_steps.internal_value_set_value (None, 'Variable', value, opts['ReturnVariable.Global'])
		if 'UpdateGUI.Value' in opts and isinstance (opts['UpdateGUI.Value'], str):
			self.cls_steps.cls_gui.internal_object_update (isolation, opts['UpdateGUI.Value'], Value=value, Trigger=(True if 'UpdateGUI.Value.Trigger' in opts and opts ['UpdateGUI.Value.Trigger'] is True else False))

		return True, {'Status':'OK', 'Data':value}
	
	
	def step_variable_set (self, isolation, **opts):
		if 'Multiple' in opts and isinstance (opts['Multiple'], list):
			variables = 0
			for entry in opts['Multiple']:
				if not isinstance (entry, dict):
					return False, {'Status':'WARNING', 'Title':'DATA:step_variable_set: Aborted', 'Message':'No ReturnVariable or ReturnVariable.Global was provided in Multiple.'}
				if 'ReturnVariable' not in entry and 'ReturnVariable.Global' not in entry:
					return False, {'Status':'WARNING', 'Title':'DATA:step_variable_set: Aborted', 'Message':'No ReturnVariable or ReturnVariable.Global was provided in Multiple.'}
				if 'Value' not in entry:
					return False, {'Status':'WARNING', 'Title':'DATA:step_variable_set: Aborted', 'Message':'No Value was provided in Multiple.'}
				value = entry['Value']
				if 'Convert' in entry and isinstance (entry['Convert'], list):
					for convert in entry['Convert']:
						status, value = datahandling_convert (value, **convert)
					if status is not True:
						return status, value
				if 'ReturnVariable' in entry and isinstance (entry['ReturnVariable'], str):
					variables += 1
					self.cls_steps.internal_value_set_value (isolation, 'Variable', value, entry['ReturnVariable'])
				if 'ReturnVariable.Global' in entry and isinstance (entry['ReturnVariable.Global'], str):
					variables += 1
					self.cls_steps.internal_value_set_value (None, 'Variable', value, entry['ReturnVariable.Global'])
			return True, {'Status':'OK', 'Message':str (variables) + ' variables updated.'}
		
		if 'ReturnVariable' not in opts and 'ReturnVariable.Global' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_variable_set: Aborted', 'Message':'No ReturnVariable or ReturnVariable.Global was provided.'}
		if 'Value' in opts:
			value = opts['Value']
		elif 'GUI.Clipboard' in opts and opts['GUI.Clipboard'] is True:
			success = False
			for x in range (0, 20):
				try:
					value = self.cls_steps.cls_gui.objects['Form'].clipboard_get ()
					success = True
					break
				except Exception as e:
					time.sleep (0.1)
			if success is False:
				return False, {'Status':'WARNING', 'Title':'DATA:step_variable_set: Aborted', 'Message':'Clipboard could not be copied.\nError=' + str (e)}
		else:
			return False, {'Status':'WARNING', 'Title':'DATA:step_variable_set: Aborted', 'Message':'No logic for which value to set was provided.'}
		
		if 'Convert' in opts and isinstance (opts['Convert'], list):
			for convert in opts['Convert']:
				status, value = datahandling_convert (value, **convert)
			if status is not True:
				return status, value
		
		fields = []
		if 'ReturnVariable' in opts and isinstance (opts['ReturnVariable'], str):
			fields.append ('Variable "' + str (opts['ReturnVariable']) + '"')
			self.cls_steps.internal_value_set_value (isolation, 'Variable', value, opts['ReturnVariable'])
		if 'ReturnVariable.Global' in opts and isinstance (opts['ReturnVariable.Global'], str):
			fields.append ('GlobalVariable "' + str (opts['ReturnVariable.Global']) + '"')
			self.cls_steps.internal_value_set_value (None, 'Variable', value, opts['ReturnVariable.Global'])
		
		return True, {'Status':'OK', 'Message':' and '.join (fields) + ' was updated with "' + (str (value) if len (str (value)) < 100 else str (value)[:100] + '...') + '".'}
	
	
	def step_variable_change (self, isolation, **opts):
		if 'Variable' not in opts and 'Variable.Global' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'No Variable or Variable.Global was provided.'}
		if 'Logic' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'No Logic was provided.'}
		if 'Value' not in opts and (opts['Logic'].endswith ('.Add') or opts['Logic'].endswith ('.Remove')):
			return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'No Value was provided.'}
		if 'Variable' in opts and not isinstance (opts['Variable'], str):
			return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'The value of Variable must be a string.'}
		if 'Variable.Global' in opts and not isinstance (opts['Variable.Global'], str):
			return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'The value of Variable.Global must be a string.'}
		
		with self.objects['Lock']['Function.Variable.Change']:
			func_isolation = (isolation if 'Variable' in opts else None)
			func_variable = (opts['Variable'] if 'Variable' in opts else opts['Variable.Global'])
			if self.cls_steps.internal_value_exists (func_isolation, 'Variable', func_variable) is True:
				variable = self.cls_steps.internal_value_get_value (func_isolation, 'Variable', func_variable)
			else:
				return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'The variable "' + str (func_variable) + '" was not found.'}
			
			if opts['Logic'].startswith ('List.') and not isinstance (variable, list):
				return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'The variable "' + str (func_variable) + '" was not a list.'}
			elif opts['Logic'].startswith ('Dict.') and not isinstance (variable, dict):
				return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'The variable "' + str (func_variable) + '" was not a dict.'}
			elif opts['Logic'].startswith ('Numeric.') and variable.__class__.__name__ not in ['int', 'float', 'Decimal']:
				return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'The variable "' + str (func_variable) + '" was not a number (type: ' + str (variable.__class__.__name__) + ').'}
			
			if opts['Logic'] == 'List.Add':
				variable.append (opts['Value'])
			elif opts['Logic'] == 'Numeric.Add':
				variable += opts['Value']
			elif opts['Logic'] == 'Numeric.Remove':
				variable -= opts['Value']
			else:
				return False, {'Status':'WARNING', 'Title':'DATA:step_variable_change: Aborted', 'Message':'The logic "' + str (opts['Logic']) + '" is unknown.'}
			
			self.cls_steps.internal_value_set_value (func_isolation, 'Variable', variable, func_variable)
			if 'UpdateGUI.Value' in opts and isinstance (opts['UpdateGUI.Value'], str):
				self.cls_steps.cls_gui.internal_object_update (isolation, opts['UpdateGUI.Value'], Value=variable, Trigger=(True if 'UpdateGUI.Value.Trigger' in opts and opts ['UpdateGUI.Value.Trigger'] is True else False))
		return True, {'Status':'OK'}

	
	def step_date (self, isolation, **opts):
		if 'Logic' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_date: Aborted', 'Message':'No Logic was provided.'}
		format = (opts['Format'] if 'Format' in opts else '%Y-%m-%d')
		
		if 'Date' in opts:
			status, result = datahandling_verify (opts['Date'], Type='Date')
			if status is not True:
				return False, {'Status':'WARNING', 'Title':'DATA:step_date: Aborted', 'Message':'The provided date "' + str (opts['Date']) + '" failed the validation with the error: ' + str (result)}
			time = datetime.datetime (int (opts['Date'][0:4]), int (opts['Date'][5:7]), int (opts['Date'][8:10]), 12, 0, 0).timestamp ()
		else:
			time = internal_zyd_time (Decimals=0)
		
		if opts['Logic'] == 'Current':
			date = internal_zyd_date (Time=time, Format=format)
		elif opts['Logic'] == 'PreviousDay':
			date = internal_zyd_date (Time=time - 86400, Format=format)
		elif opts['Logic'] == 'NextDay':
			date = internal_zyd_date (Time=time + 86400, Format=format)
		elif opts['Logic'] in ['PreviousWorkday', 'NextWorkday']:
			for x in range (1, 10):
				time += (86400 if opts['Logic'].startswith ('Next') else -86400)
				if self.internal_step_date_isworkday (time, **opts):
					date = internal_zyd_date (Time=time, Format=format)
					break
		elif opts['Logic'] in ['PreviousDay.MatchKey', 'PreviousWorkday.MatchKey', 'FirstDay.MatchKey', 'NextDay.MatchKey', 'NextWorkday.MatchKey', 'LastWorkday.MatchKey', 'FirstWorkday.MatchKey', 'LastDay.MatchKey']:
			if 'MatchKey' not in opts:
				return False, {'Status':'WARNING', 'Title':'DATA:step_date: Aborted', 'Message':'The Logic "' + str (opts['Logic']) + '" requires a MatchKey to be provided..'}
			matchkey = internal_zyd_date (Time=time, Format=opts['MatchKey'])
			for x in range (0, 10000):
				date = internal_zyd_date (Time=time, Format=format)
				time_current = copy.deepcopy (time)
				time += (86400 if (opts['Logic'].startswith ('Next') or opts['Logic'].startswith ('Last')) else -86400)
				if matchkey != internal_zyd_date (Time=time, Format=opts['MatchKey']):
					break
				elif opts['Logic'].startswith ('First') or opts['Logic'].startswith ('Last'):
					if 'Workday' in opts['Logic'] and self.internal_step_date_isworkday (time_current, **opts) and not self.internal_step_date_isworkday (time, **opts):
						break
				elif 'Workday' in opts['Logic'] and x > 0 and self.internal_step_date_isworkday (time_current, **opts):
					break
				elif 'Day' in opts['Logic'] and x > 0:
					break
		elif opts['Logic'] in ['NextMonth', 'PreviousMonth']:
			month = internal_zyd_date (Time=time, Format='%m')
			for x in range (0, 100):
				time += (86400 if opts['Logic'].startswith ('Next') else -86400)
				if month != internal_zyd_date (Time=time, Format='%m'):
					date = internal_zyd_date (Time=time, Format=format)
					break
		else:
			return False, {'Status':'WARNING', 'Title':'DATA:step_date: Aborted', 'Message':'The Logic "' + str (opts['Logic']) + '" is unknown.'}
		
		if 'ReturnVariable' in opts and isinstance (opts['ReturnVariable'], str):
			self.cls_steps.internal_value_set_value (isolation, 'Variable', date, opts['ReturnVariable'])
		if 'ReturnVariable.Global' in opts and isinstance (opts['ReturnVariable.Global'], str):
			self.cls_steps.internal_value_set_value (None, 'Variable', date, opts['ReturnVariable.Global'])
		if 'UpdateGUI.Value' in opts and isinstance (opts['UpdateGUI.Value'], str):
			self.cls_steps.cls_gui.internal_object_update (isolation, opts['UpdateGUI.Value'], Value=date, Trigger=(True if 'UpdateGUI.Value.Trigger' in opts and opts ['UpdateGUI.Value.Trigger'] is True else False))
		return True, {'Status':'OK', 'Data':date}
	
	
	def internal_step_date_isworkday (self, time, **opts):
		if int (internal_zyd_date (Time=time, Format='%u')) <= 5:
			if 'BankHolidays' in opts and isinstance (opts['BankHolidays'], list) and internal_zyd_date (Time=time, Format='%Y-%m-%d') in opts['BankHolidays']:
				return False
			return True
		return False
	
	
	def step_transform (self, isolation, **opts):
		if 'Value' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_transform: Aborted', 'Message':'No Value was provided.'}
		if 'Logic' not in opts:
			return False, {'Status':'WARNING', 'Title':'DATA:step_transform: Aborted', 'Message':'No Logic was provided.'}
		
		if opts['Logic'] == 'KeyList':
			if isinstance (opts['Value'], dict):
				value = list (opts['Value'].keys ())
			elif isinstance (opts['Value'], list):
				value = list (range (0, len (opts['Value'])))
			else:
				return False, {'Status':'WARNING', 'Title':'DATA:step_transform: Aborted', 'Message':'A dict or list is required for the "KeyList" logic.'}
		elif opts['Logic'] == 'Count':
			if isinstance (opts['Value'], (dict, list)):
				value = len (opts['Value'])
			else:
				return False, {'Status':'WARNING', 'Title':'DATA:step_transform: Aborted', 'Message':'A dict or list is required for the "Count" logic.'}
		elif opts['Logic'] == 'Regex.Findall' and (('Regex' in opts and isinstance (opts['Regex'], str)) or ('Regex.List' in opts and isinstance (opts['Regex.List'], list))):
			value = []
			for regex in (opts['Regex.List'] if 'Regex.List' in opts and isinstance (opts['Regex.List'], list) else [opts['Regex']]):
				matches = re.findall (regex, (opts['Value'].decode ('UTF-8') if isinstance (opts['Value'], bytes) else opts['Value']))
				if len (matches) > 0:
					for match in matches:
						if isinstance (match, tuple):
							match = list (match)
						if match not in value:
							value.append (match)
		elif (opts['Logic'] == 'List.Modify' and isinstance (opts['Value'], list)) or (opts['Logic'] == 'Dict.Modify' and isinstance (opts['Value'], dict)):
			value = copy.deepcopy (opts['Value'])
			for key in (range (0, len (value)) if isinstance (opts['Value'], list) else value.keys ()):
				if isinstance (value[key], str):
					if 'Modify.Prefix' in opts and isinstance (opts['Modify.Prefix'], str):
						value[key] = opts['Modify.Prefix'] + value[key]
					if 'Modify.Suffix' in opts and isinstance (opts['Modify.Suffix'], str):
						value[key] += opts['Modify.Suffix']
					if 'Modify.Replace' in opts and isinstance (opts['Modify.Replace'], dict):
						for rkey in opts['Modify.Replace'].keys ():
							if str (rkey) in value[key]:
								value[key] = value[key].replace (str (rkey), str (opts['Modify.Replace'][rkey]))
				value[key] = value[key]
		elif opts['Logic'] == '2DList.ToDict' and isinstance (opts['Value'], list):
			value = {}
			for entry in opts['Value']:
				if not isinstance (entry, list) or len (entry) != 2:
					return False, {'Status':'WARNING', 'Title':'DATA:step_transform: Aborted', 'Message':'The logic "' + str (opts['Logic']) + '" got a list which didn\'t only contain a 2D list with 2 entries.'}
				value[entry[0]] = entry[1]
		elif opts['Logic'] == 'Dict.ModifyKey' and isinstance (opts['Value'], dict):
			value = {}
			for key in opts['Value'].keys ():
				newkey = copy.deepcopy (key)
				if 'Modify.Prefix' in opts and isinstance (opts['Modify.Prefix'], str):
					newkey = opts['Modify.Prefix'] + newkey
				if 'Modify.Suffix' in opts and isinstance (opts['Modify.Suffix'], str):
					newkey = newkey + opts['Modify.Suffix']
				if 'Modify.Replace' in opts and isinstance (opts['Modify.Replace'], dict):
					for rkey in opts['Modify.Replace'].keys ():
						if str (rkey) in newkey:
							newkey = newkey.replace (str (rkey), str (opts['Modify.Replace'][rkey]))
				value[newkey] = opts['Value'][key]
		elif opts['Logic'] == 'String.Modify' and isinstance (opts['Value'], str):
			if 'Modify.Prefix' in opts and isinstance (opts['Modify.Prefix'], str):
				value = opts['Modify.Prefix'] + opts['Value']
			if 'Modify.Suffix' in opts and isinstance (opts['Modify.Suffix'], str):
				value = opts['Value'] + opts['Modify.Suffix']
			if 'Modify.Replace' in opts and isinstance (opts['Modify.Replace'], dict):
				value = copy.deepcopy (opts['Value'])
				for rkey in opts['Modify.Replace'].keys ():
					if str (rkey) in newkey:
						value = value.replace (str (rkey), str (opts['Modify.Replace'][rkey]))
		elif opts['Logic'] == 'ArrayToVariables' and 'ArrayToVariables' in opts and isinstance (opts['ArrayToVariables'], list):
			for variable in opts['ArrayToVariables']:
				if not isinstance (variable, dict) or 'Keys' not in variable or not isinstance (variable['Keys'], list):
					return False, {'Status':'WARNING', 'Title':'DATA:step_transform: Aborted', 'Message':'The variable "' + str (variable) + '" is not valid.'}
				temp = opts['Value']
				for key in variable['Keys']:
					if (isinstance (temp, list) and isinstance (key, int) and key < len (temp)) or (isinstance (temp, dict) and key in temp):
						temp = temp[key]
					else:
						return False, {'Status':'WARNING', 'Title':'DATA:step_transform: Aborted', 'Message':'The key "' + str (key) + '" was not found in the key chain "' + str (variable['Keys']) + '".'}
				if 'ReturnVariable' in variable and isinstance (variable['ReturnVariable'], str):
					self.cls_steps.internal_value_set_value (isolation, 'Variable', temp, variable['ReturnVariable'])
				if 'ReturnVariable.Global' in variable and isinstance (variable['ReturnVariable.Global'], str):
					self.cls_steps.internal_value_set_value (None, 'Variable', temp, variable['ReturnVariable.Global'])
				if 'UpdateGUI.Value' in variable and isinstance (variable['UpdateGUI.Value'], str):
					self.cls_steps.cls_gui.internal_object_update (isolation, variable['UpdateGUI.Value'], Value=temp, Trigger=(True if 'UpdateGUI.Value.Trigger' in variable and variable['UpdateGUI.Value.Trigger'] is True else False))
		elif opts['Logic'] in ['ArrayDiff', 'ArrayDiff.Matched', 'ArrayDiff.Mismatched'] and 'ArrayDiff' in opts and isinstance (opts['ArrayDiff'], (list, dict)):
			value = ([] if isinstance (opts['Value'], list) else {})
			if isinstance (opts['Value'], list):
				if opts['Logic'] == 'ArrayDiff.Matched':
					for svalue in opts['Value']:
						if svalue in opts['ArrayDiff']:
							value.append (svalue)
				elif opts['Logic'] == 'ArrayDiff.Mismatched' or opts['Logic'] == 'ArrayDiff':
					for svalue in opts['Value']:
						if svalue not in opts['ArrayDiff']:
							value.append (svalue)
					if opts['Logic'] == 'ArrayDiff.Mismatched':
						for svalue in opts['ArrayDiff']:
							if svalue not in opts['Value'] and svalue not in value:
								value.append (svalue)
			elif isinstance (opts['Value'], dict):
				if opts['Logic'] == 'ArrayDiff.Matched':
					for key in opts['Value'].keys ():
						if key in opts['ArrayDiff'] and opts['Value'][key] == opts['ArrayDiff'][key]:
							value[key] = opts['Value'][key]
				elif opts['Logic'] == 'ArrayDiff.Mismatched' or opts['Logic'] == 'ArrayDiff':
					for key in opts['Value'].keys ():
						if key not in opts['ArrayDiff'] or opts['Value'][key] != opts['ArrayDiff'][key]:
							value[key] = opts['Value'][key]
					if opts['Logic'] == 'ArrayDiff.Mismatched':
						for key in opts['ArrayDiff'].keys ():
							if key not in opts['Value'] or opts['ArrayDiff'][key] != opts['Value'][key]:
								value[key] = opts['ArrayDiff'][key]
		else:
			return False, {'Status':'WARNING', 'Title':'DATA:step_transform: Aborted', 'Message':'The logic "' + str (opts['Logic']) + '" is not supported.'}
		
		if 'ReturnVariable' in opts and isinstance (opts['ReturnVariable'], str):
			self.cls_steps.internal_value_set_value (isolation, 'Variable', value, opts['ReturnVariable'])
		if 'ReturnVariable.Global' in opts and isinstance (opts['ReturnVariable.Global'], str):
			self.cls_steps.internal_value_set_value (None, 'Variable', value, opts['ReturnVariable.Global'])
		if 'UpdateGUI.Value' in opts and isinstance (opts['UpdateGUI.Value'], str):
			self.cls_steps.cls_gui.internal_object_update (isolation, opts['UpdateGUI.Value'], Value=value, Trigger=(True if 'UpdateGUI.Value.Trigger' in opts and opts ['UpdateGUI.Value.Trigger'] is True else False))

		return True, {'Status':'OK'}
	
	
	def step_convert (self, isolation, **opts):
		debuglog = (True if 'DebugLog' in opts and opts['DebugLog'] is True else False)
		debuglogdetailed = (True if 'DebugLog.Detailed' in opts and opts['DebugLog.Detailed'] is True else False)
		if 'Value' in opts and 'Convert' in opts and isinstance (opts['Convert'], list):
			value = copy.deepcopy (opts['Value'])
			for convert in opts['Convert']:
				status, value = datahandling_convert (value, **convert)
				if debuglogdetailed:
					self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_convert', 'Convert':convert, 'Input':opts['Value'], 'Output':value})
				if status is not True:
					return status, value
			if debuglog:
				self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_convert', 'Input':opts['Value'], 'Output':value})
			
			if 'ReturnVariable' in opts and isinstance (opts['ReturnVariable'], str):
				self.cls_steps.internal_value_set_value (isolation, 'Variable', value, opts['ReturnVariable'])
			if 'ReturnVariable.Global' in opts and isinstance (opts['ReturnVariable.Global'], str):
				self.cls_steps.internal_value_set_value (None, 'Variable', value, opts['ReturnVariable.Global'])
			if 'UpdateGUI.Value' in opts and isinstance (opts['UpdateGUI.Value'], str):
				self.cls_steps.cls_gui.internal_object_update (isolation, opts['UpdateGUI.Value'], Value=value, Trigger=(True if 'UpdateGUI.Value.Trigger' in opts and opts ['UpdateGUI.Value.Trigger'] is True else False))
			return True, {'Status':'OK', 'Data':value}
		elif 'Value.List2D' in opts and 'Convert.List2D' in opts and isinstance (opts['Value.List2D'], list) and isinstance (opts['Value.List2D'][0], list) and isinstance (opts['Convert.List2D'], list):
			colsconvert = len (opts['Convert.List2D'])
			if len (opts['Value.List2D'][0]) == colsconvert:
				data = copy.deepcopy (opts['Value.List2D'])
				for irow in range (0, len (data)):
					cols = len (data[irow])
					if cols != colsconvert:
						return False, {'Status':'WARNING', 'Function':'step_convert', 'Title':'DATA:step_convert: Aborted', 'Message':'List data doesn\'t have the same amount of columns as the conversion rules (line=' + str (irow) + ' with ' + str (cols) + ' and convert ' + str (colsconvert) + ' columns).'}
					for icol in range (0, cols):
						for convert in opts['Convert.List2D'][icol]:
							status, data[irow][icol] = datahandling_convert (data[irow][icol], **convert)
							if debuglogdetailed:
								self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_convert', 'LineNo':irow, 'ColumnNo':icol, 'Convert':convert, 'Input':opts['Value.List2D'][irow][icol], 'Output':data[irow][icol]})
							if status is not True:
								return status, data[irow][icol]
						if debuglog:
							self.cls_steps.internal_console_debugging (**{'Class':'DATA', 'Function':'step_convert', 'LineNo':irow, 'Input':opts['Value.List2D'][irow], 'Output':data[irow]})
			else:
				return False, {'Status':'WARNING', 'Function':'step_convert', 'Title':'DATA:step_convert: Aborted', 'Message':'List data doesn\'t have the same amount of columns as the conversion rules'}
			if 'ReturnVariable' in opts and isinstance (opts['ReturnVariable'], str):
				self.cls_steps.internal_value_set_value (isolation, 'Variable', data, opts['ReturnVariable'])
			if 'ReturnVariable.Global' in opts and isinstance (opts['ReturnVariable.Global'], str):
				self.cls_steps.internal_value_set_value (None, 'Variable', data, opts['ReturnVariable.Global'])
			return True, {'Status':'OK', 'Data':data}
		else:
			return False, {'Status':'WARNING', 'Function':'step_convert', 'Title':'DATA:step_convert: Aborted', 'Message':'Unknown opts=' + str (opts)}
		return True, {'Status':'ERROR', 'Message':'Unknown...'}
	
	
	def step_replace (self, isolation, **opts):
		value = copy.deepcopy (opts['Value'])
		if 'Replace.Dict' in opts and isinstance (opts['Replace.Dict'], dict):
			for key in opts['Replace.Dict'].keys ():
				if str (key) in value:
					value = value.replace (str (key), str (opts['Replace.Dict'][key]))
		if 'Replace.Dict.Regex' in opts and isinstance (opts['Replace.Dict.Regex'], dict):
			for key in opts['Replace.Dict.Regex'].keys ():
				value = re.sub (key, opts['Replace.Dict.Regex'][key], value)
		
		if 'ReturnVariable' in opts and isinstance (opts['ReturnVariable'], str):
			self.cls_steps.internal_value_set_value (isolation, 'Variable', value, opts['ReturnVariable'])
		if 'ReturnVariable.Global' in opts and isinstance (opts['ReturnVariable.Global'], str):
			self.cls_steps.internal_value_set_value (None, 'Variable', value, opts['ReturnVariable.Global'])
		if 'UpdateGUI.Value' in opts and isinstance (opts['UpdateGUI.Value'], str):
			self.cls_steps.cls_gui.internal_object_update (isolation, opts['UpdateGUI.Value'], Value=value, Trigger=(True if 'UpdateGUI.Value.Trigger' in opts and opts ['UpdateGUI.Value.Trigger'] is True else False))
		return True, {'Status':'OK', 'Data':value}
	
	
	
	def internal_read_file (self, **opts):
		data = []
		if 'Type' in opts and opts['Type'] == 'CSV':
			with open (opts['File'], newline=(opts['CSV.Newline'] if 'CSV.Newline' in opts else '\n')) as csvfile:
				csvreader = csv.reader (csvfile, delimiter=(opts['CSV.Delimiter'] if 'CSV.Delimiter' in opts else ','), quotechar=(opts['CSV.Quotechar'] if 'CSV.Quotechar' in opts else '"'))
				for row in csvreader:
					data.append (list (row))
		elif 'Type' in opts and opts['Type'] == 'JSON':
			with open (opts['File']) as file:
				json = file.read (os.path.getsize (opts['File']))
				status, data = zyd_json (json, Decode=True)
		elif 'Type' in opts and opts['Type'] == 'Binary':
			with open (opts['File'], 'rb') as file:
				data = file.read ()
		elif opts['Format'] == 'XLSX':
			workbook = openpyxl.load_workbook (filename=opts['File'], read_only=True)
			if 'EXCEL.FirstSheet' in opts and opts['EXCEL.FirstSheet']:
				sheet = workbook[list (workbook.sheetnames)[0]]
				data = []
				for row in sheet.iter_rows ():
					returnrow = []
					for value in row:
						returnrow.append ((self.cls_steps.cls_excel.internal_read_file_openpyxl_sheet_respect_formatting (value.value, value.number_format) if 'EXCEL.RespectFormatting' in opts and opts['EXCEL.RespectFormatting'] is True else value.value))
					data.append (returnrow)
			workbook.close ()
		
		return data
	
	
	def internal_convert_list_data (self, data):
		if isinstance (data, list):
			for key in range (0, len (data)):
				data[key] = self.internal_convert_list_data (data[key])
		elif isinstance (data, dict):
			if 'Convert' in data and 'Value' in data and isinstance (data['Convert'], list):
				value = copy.deepcopy (data['Value'])
				for convert in data['Convert']:
					status, value = datahandling_convert (value, **convert)
					if status is not True:
						print ('internal_convert_list_data.BREAK::' + str (data) + '::' + str (value))
						return data['Value']
				return value
		return data
	