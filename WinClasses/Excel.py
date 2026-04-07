# Copyright (c) 2025-2026 Erik Persson
# Licensed for free use only. Copying, modification, or redistribution prohibited.

import builtins
from logging import exception, error
from Functions.Common import *
from Functions.DataHandling import *
for name in dir ():
	if not name.startswith ("_"):
		builtins.__dict__[name] = globals ()[name]
import openpyxl
builtins.openpyxl = openpyxl
import openpyxl.styles
builtins.openpyxl.styles = openpyxl.styles
import xlrd
builtins.xlrd = xlrd
import re
builtins.re = re


class Excel:
	def __init__ (self, cls_steps):
		self.cls_steps = cls_steps
		self.objects = self.cls_steps.objects
		
	
	def dynamic_function (self, function, to_class = True):
		if to_class is True:
			func = {}
			exec (function['Code'], func)
			setattr (Excel, function['Function'], func[function['Function']])
		else:
			exec (function['Code'], globals ())
			builtins.__dict__[function['Function']] = globals ()[function['Function']]
	
	
	def execute (self, isolation = None, **opts):
		try:
			if opts['Function'] == 'EXCEL:Open':
				return self.step_open (isolation, **opts)
			elif opts['Function'] == 'EXCEL:Close':
				return self.step_close (isolation, **opts)
			elif opts['Function'] == 'EXCEL:Save':
				return self.step_save (isolation, **opts)
			elif opts['Function'] == 'EXCEL:Change':
				return self.step_change (isolation, **opts)
			elif opts['Function'] == 'EXCEL:Read':
				return self.step_read (isolation, **opts)
			else:
				return False, {'Status':'WARNING', 'Title':'EXCEL:execute: Aborted', 'Message':'Unknown execute:' + str (opts)}
		except Exception as e:
			st = internal_zyd_stacktrace (e, FunctionOpts=opts)
			return False, {'Status':'CRASH', 'Title':'EXCEL:execute: Crashed', 'Message':'Crashed with the error: ' + str (e) + ' for step:\n' + str (opts), 'StackTrace':st}

	
	def step_open (self, isolation, **opts):
		if 'ID' not in opts:
			return False, {'Status':'WARNING', 'Title':'EXCEL:open: Aborted', 'Message':'No ID was provided for the function.'}
		if opts['ID'] not in self.objects['Excel']:
			self.objects['Excel'][opts['ID']] = {
				'Object':openpyxl.load_workbook (filename=opts['File']),
				'File':opts['File']
			}
			return True, {'Status':'OK'}
		return False, {'Status':'WARNING', 'Title':'EXCEL:open: Aborted', 'Message':'The ID "' + str (opts['ID']) + '" is already in use.'}
	
	
	def step_close (self, isolation, **opts):
		if 'ID' not in opts:
			return False, {'Status':'WARNING', 'Title':'EXCEL:close: Aborted', 'Message':'No ID was provided for the function.'}
		if opts['ID'] in self.objects['Excel']:
			self.objects['Excel'][opts['ID']]['Object'].close ()
			del (self.objects['Excel'][opts['ID']])
			return True, {'Status':'OK'}
		return False, {'Status':'WARNING', 'Title':'EXCEL:close: Aborted', 'Message':'No Excel object was found for ID "' + str (opts['ID']) + '".'}
	
	
	def step_save (self, isolation, **opts):
		if 'ID' not in opts:
			return False, {'Status':'WARNING', 'Title':'EXCEL:save: Aborted', 'Message':'No ID was provided for the function.'}
		if opts['ID'] in self.objects['Excel']:
			file = (opts['File'] if 'File' in opts else self.objects['Excel'][opts['ID']]['File'])
			if 'ErrorIfOpen' in opts and opts['ErrorIfOpen'] is True:
				filesplit = os.path.split (file)
				if os.path.exists (filesplit[0] + '\\~$' + filesplit[1]):
					return False, {'Status':'ERROR', 'Title':'EXCEL:save: Aborted', 'Message':'The excel file is already open in Excel.'}
				elif os.path.exists (filesplit[0] + '\\.~lock.' + filesplit[1] + '#'):
					return False, {'Status':'ERROR', 'Title':'EXCEL:save: Aborted', 'Message':'The excel file is already open in Libre calc.'}
			return True, {'Status':'OK'}
		return False, {'Status':'WARNING', 'Title':'EXCEL:save: Aborted', 'Message':'No Excel object was found for ID "' + str (opts['ID']) + '".'}
	
	
	def step_change (self, isolation, **opts):
		if 'ID' not in opts:
			return False, {'Status':'WARNING', 'Title':'EXCEL:change: Aborted', 'Message':'No ID was provided for the function.'}
		if 'Sheet' not in opts:
			return False, {'Status':'WARNING', 'Title':'EXCEL:change: Aborted', 'Message':'No Sheet was provided for the function.'}
		if opts['ID'] not in self.objects['Excel']:
			return False, {'Status':'WARNING', 'Title':'EXCEL:change: Aborted', 'Message':'No Excel object was found for ID "' + str (opts['ID']) + '".'}
		if opts['Sheet'] not in self.objects['Excel'][opts['ID']]['Object']:
			return False, {'Status':'WARNING', 'Title':'EXCEL:change: Aborted', 'Message':'No Excel sheet was found for sheet "' + str (opts['Sheet']) + '".'}
		
		styleformat = None
		if 'ValueFormat' in opts:
			if opts['ValueFormat'] == 'Date':
				if 'Omoikane ISO Date' not in self.objects['Excel'][opts['ID']]['Object'].named_styles:
					styleformat = openpyxl.styles.NamedStyle (name='Omoikane ISO Date', number_format='YYYY-MM-DD')
				else:
					styleformat = 'Omoikane ISO Date'
			else:
				return False, {'Status':'WARNING', 'Title':'EXCEL:change: Aborted', 'Message':'The ValueFormat "' + str (opts['ValueFormat']) + '" is unknown.'}
				
		
		if 'Cell' in opts and isinstance (opts['Cell'], str):
			self.objects['Excel'][opts['ID']]['Object'][opts['Sheet']][opts['Cell']].value = opts['Value']
			if styleformat is not None:
				self.objects['Excel'][opts['ID']]['Object'][opts['Sheet']][opts['Cell']].style = styleformat
		return True, {'Status':'OK'}
	
	
	def step_read (self, isolation, **opts):
		if 'ID' not in opts:
			return False, {'Status':'WARNING', 'Title':'EXCEL:read: Aborted', 'Message':'No ID was provided for the function.'}
		if 'Sheet' not in opts:
			return False, {'Status':'WARNING', 'Title':'EXCEL:read: Aborted', 'Message':'No Sheet was provided for the function.'}
		if opts['ID'] not in self.objects['Excel']:
			return False, {'Status':'WARNING', 'Title':'EXCEL:read: Aborted', 'Message':'No Excel object was found for ID "' + str (opts['ID']) + '".'}
		if opts['Sheet'] not in self.objects['Excel'][opts['ID']]['Object']:
			return False, {'Status':'WARNING', 'Title':'EXCEL:read: Aborted', 'Message':'No Excel sheet was found for sheet "' + str (opts['Sheet']) + '".'}
		
		if 'Cell' in opts and isinstance (opts['Cell'], str):
			data = (self.internal_read_file_openpyxl_sheet_respect_formatting (self.objects['Excel'][opts['ID']]['Object'][opts['Sheet']][opts['Cell']].value, self.objects['Excel'][opts['ID']]['Object'][opts['Sheet']][opts['Cell']].number_format) if 'RespectFormatting' in opts and opts['RespectFormatting'] is True else self.objects['Excel'][opts['ID']]['Object'][opts['Sheet']][opts['Cell']].value)
		elif 'Row' in opts and isinstance (opts['Row'], int):
			data = []
			for cell in self.objects['Excel'][opts['ID']]['Object'][opts['Sheet']][opts['Row']]:
				data.append ((self.internal_read_file_openpyxl_sheet_respect_formatting (cell.value, cell.number_format) if 'RespectFormatting' in opts and opts['RespectFormatting'] is True else cell.value))
		elif 'Complete' in opts and opts['Complete'] is True:
			data = []
			for row in range (1, self.objects['Excel'][opts['ID']]['Object'][opts['Sheet']].max_row + 1):
				line = []
				for column in range (1, self.objects['Excel'][opts['ID']]['Object'][opts['Sheet']].max_column + 1):
					cell = self.objects['Excel'][opts['ID']]['Object'][opts['Sheet']].cell (row=row, column=column)
					line.append ((self.internal_read_file_openpyxl_sheet_respect_formatting (cell.value, cell.number_format) if 'RespectFormatting' in opts and opts['RespectFormatting'] is True else cell.value))
				data.append (line)
		
		if 'ReturnVariable' in opts and isinstance (opts['ReturnVariable'], str):
			self.cls_steps.internal_value_set_value (isolation, 'Variable', data, opts['ReturnVariable'])
		if 'ReturnVariable.Global' in opts and isinstance (opts['ReturnVariable.Global'], str):
			self.cls_steps.internal_value_set_value (None, 'Variable', data, opts['ReturnVariable.Global'])
		return True, {'Status':'OK', 'Data':data}


	def internal_read_file_openpyxl_sheet_respect_formatting (self, value, formatting):
		if not isinstance (formatting, str):
			return value
		if formatting[:5] == '#,##0':
#			"""	Covers formatting like
#				#,##0.000" PCE";\-#,##0.000" PCE";#,##0.000" PCE"
#				#,##0" PCE";\-#,##0" PCE";#,##0" PCE"
#				#,##0.000" PLN";\-#,##0.000" PLN";#,##0.000" PLN"
#				#,##0.00" EUR";\-#,##0.00" EUR";#,##0.00" EUR"
#				#,##0.000" KG";\-#,##0.000" KG";#,##0.000" KG"
#			"""
			regex = re.search (r'^#,##0[0\.]*" [A-Z]+"', formatting)
			if regex is not None:
				regex2 = re.search (r'" [A-Z]+"$', formatting[regex.span ()[0]:regex.span ()[1]])
				if regex2 is not None:
					value = str (value) + str (regex2.group ()[1:-1])
		elif ' "#,##0' in formatting:
			regex = re.search (r'^"[$€£] "#,##0[0\.]*', formatting)
			if regex is not None:
				regex2 = re.search ('^"[$€£] "', formatting[regex.span ()[0]:regex.span ()[1]])
				if regex2 is not None:
					value = str (regex2.group ()[1:-1]) + str (value)
		return value